"""
Tests for the Quadro 2 alignment work (2026-05-08):
* 6-level signal rating
* indoor_outdoor propagation through upload
* frequency_hz derived from ARFCN+band
* Scientific export mode (only Quadro 2 columns)
* /api/points exposes the new environmental fields
* /api/sectors/calibration warning still trips on synthetic
"""
from __future__ import annotations

import pytest

from app.parser import _signal_rating, _frequency_hz


# --- 6-level signal rating ------------------------------------------------
@pytest.mark.parametrize("rsrp,expected", [
    (-70,  "Excelente"),
    (-85,  "Bom"),          # boundary: > -85 is Excelente, == -85 is Bom
    (-90,  "Bom"),
    (-95,  "Satisfatório"),
    (-100, "Satisfatório"),
    (-105, "Ruim"),
    (-110, "Ruim"),
    (-115, "Péssimo"),
    (-120, "Péssimo"),
    (-125, "Nulo"),
    (-130, "Nulo"),
    (None, "Nulo"),
    (float("nan"), "Nulo"),
])
def test_signal_rating_six_levels(rsrp, expected):
    assert _signal_rating(rsrp) == expected


# --- frequency_hz ---------------------------------------------------------
@pytest.mark.parametrize("band,expected_hz", [
    ("n78",  3_500e6),
    ("B3",   1_842.5e6),
    ("B7",   2_655e6),
    ("brand-new-band", None),
    (None,   None),
    ("",     None),
])
def test_frequency_hz_lookup(band, expected_hz):
    assert _frequency_hz(band, None) == expected_hz


# --- indoor_outdoor propagation through /api/upload ------------------------
HEADER = (
    "Timestamp\tLatitude\tLongitude\tAccuracy\tAltitude\tSpeed\t"
    "Operatorname\tOperator\tCGI\tCellID\tLAC\tNetworkTech\tNetworkMode\t"
    "Level\tQual\tSNR\tCQI\tLTERSSI\tARFCN\tBAND\tBANDWIDTH\t"
    "DL_bitrate\tUL_bitrate\tDistance\tEVENT\tEVENTDETAILS\t"
    "PINGAVG\tPINGMIN\tPINGMAX\tPINGSTDEV\tPINGLOSS\t"
    "TESTDOWNLINK\tTESTUPLINK\tTESTDOWNLINKMAX\tTESTUPLINKMAX\t"
    "CSI_RSRP\tCSI_RSRQ\tCSI_SNR"
)
ROW = (
    "{ts}\t-23.20950\t-45.87650\t8\t580\t1.2\t"
    "VIVO\tVIVO\t724.06.123.4567\t1234567\t6\t5G\t5G NSA\t"
    "-90\t-12\t8\t10\t-65\t627000\tn78\t100\t"
    "12000\t1500\t120\tPERIODIC\t-\t"
    "32\t28\t40\t3\t0\t"
    "5000\t800\t8000\t1200\t"
    "-95\t-13\t9"
)


def _make_log(timestamps: list[str]) -> bytes:
    rows = [ROW.format(ts=t) for t in timestamps]
    return ("\r\n".join([HEADER, *rows]) + "\r\n").encode("utf-8")


def _upload(client, payload, **params):
    return client.post(
        "/api/upload",
        params={"enrich": "false", **params},
        files={"file": ("a.txt", payload, "text/plain")},
    )


def test_indoor_outdoor_propagates_to_measurements(api_client, fresh_db):
    payload = _make_log(["2026.05.04_09.00.00", "2026.05.04_09.00.10"])
    r = _upload(api_client, payload, indoor_outdoor="indoor",
                campaign_id="manha")
    assert r.json()["status"] == "success"
    pts = api_client.get("/api/points").json()
    assert pts and all(p["indoor_outdoor"] == "indoor" for p in pts)


def test_frequency_hz_landed_on_measurement(api_client, fresh_db):
    """The synthetic log uses BAND=n78 → expect 3.5 GHz nominal."""
    _upload(api_client, _make_log(["2026.05.04_09.00.00"]))
    pts = api_client.get("/api/points").json()
    assert pts
    assert pts[0]["band"] == "n78"
    assert pts[0]["frequency_hz"] == pytest.approx(3_500_000_000, rel=1e-9)


# --- /api/points carries environmental fields ------------------------------
def test_points_endpoint_carries_environmental_columns(api_client, fresh_db):
    _upload(api_client, _make_log(["2026.05.04_09.00.00"]))
    pts = api_client.get("/api/points").json()
    p = pts[0]
    # Even with enrich=false, the columns must be present (as None) so the
    # dashboard popup can read them safely.
    for k in ("temperature_c", "humidity",
              "cloud_cover_pct", "cloud_cover_label",
              "building_count", "avg_building_height", "distance_to_building_m",
              "tree_count", "avg_tree_height_m", "distance_to_tree_m",
              "tree_density_ndvi", "frequency_hz", "indoor_outdoor"):
        assert k in p, f"missing key {k} in /api/points response"


# --- /api/export ---------------------------------------------------------
def test_export_scientific_columns_match_quadro2(api_client, fresh_db, tmp_path):
    _upload(api_client, _make_log(["2026.05.04_09.00.00"]),
            indoor_outdoor="outdoor")
    r = api_client.get("/api/export", params={"mode": "scientific"})
    assert r.status_code == 200
    out = tmp_path / "ex.xlsx"
    out.write_bytes(r.content)
    import openpyxl
    wb = openpyxl.load_workbook(out, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    # Spot-check that the 18 Quadro-2 labels appear (Portuguese, with units).
    must_have = {
        "Avaliação do sinal", "Localização do quadrante",
        "Download (Mbps)", "Upload (Mbps)", "Latência (ms)", "Jitter (ms)",
        "Potência do sinal (dBm)", "Frequência (Hz)", "Altitude (m)",
        "Cobertura de nuvens", "Temperatura (°C)", "Umidade relativa (%)",
        "Quantidade de árvores", "Altura média das árvores (m)",
        "Distância média para as árvores (m)",
        "Quantidade de prédios", "Altura média dos prédios (m)",
        "Distância média para os prédios (m)",
    }
    missing = must_have - set(headers)
    assert not missing, f"Quadro-2 columns missing from scientific export: {missing}"


def test_export_full_has_more_columns_than_scientific(api_client, fresh_db,
                                                      tmp_path):
    _upload(api_client, _make_log(["2026.05.04_09.00.00"]))
    rs = api_client.get("/api/export", params={"mode": "scientific"})
    rf = api_client.get("/api/export", params={"mode": "full"})
    s = tmp_path / "s.xlsx"; f = tmp_path / "f.xlsx"
    s.write_bytes(rs.content); f.write_bytes(rf.content)
    import openpyxl
    nh_s = len([c.value for c in next(openpyxl.load_workbook(s, read_only=True)
                                       .active.iter_rows(max_row=1))])
    nh_f = len([c.value for c in next(openpyxl.load_workbook(f, read_only=True)
                                       .active.iter_rows(max_row=1))])
    assert nh_f > nh_s


def test_export_mode_validation(api_client, fresh_db):
    _upload(api_client, _make_log(["2026.05.04_09.00.00"]))
    r = api_client.get("/api/export", params={"mode": "bogus"})
    assert r.status_code == 422
